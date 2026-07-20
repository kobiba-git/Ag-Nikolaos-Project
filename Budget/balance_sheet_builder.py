from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY="1F3864"; BLUE="2E5496"; LBLUE="D6E0F0"; GREY="F2F2F2"; GREEN="C6EFCE"; RED="FFC7CE"; YELL="FFF2CC"; ORANGE="FCE4D6"
WHITE="FFFFFF"
def F(sz=10,b=False,color="000000"): return Font(name="Arial",size=sz,bold=b,color=color)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color="BFBFBF"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
EUR='€#,##0.00'; EUR0='€#,##0'
center=Alignment(horizontal="center",vertical="center",wrap_text=True)
left=Alignment(horizontal="left",vertical="center",wrap_text=True)
right=Alignment(horizontal="right",vertical="center")
wb=Workbook()

def hdr(ws,row,cols,fillc=BLUE):
    for j,c in enumerate(cols,1):
        cell=ws.cell(row=row,column=j,value=c); cell.font=F(10,True,WHITE); cell.fill=fill(fillc); cell.alignment=center; cell.border=border

def title(ws,text,sub,ncol):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncol)
    c=ws.cell(1,1,text); c.font=F(16,True,WHITE); c.fill=fill(NAVY); c.alignment=Alignment(horizontal="left",vertical="center")
    ws.row_dimensions[1].height=28
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=ncol)
    c=ws.cell(2,1,sub); c.font=F(9,False,"595959"); c.alignment=Alignment(horizontal="left",vertical="center")

# ============ CURRENT NBG BALANCE ANCHOR ============
# This is the ONLY place these numbers should live. Whenever Kobi/Vicky reports a
# new NBG balance (statement or spot-check), update CURRENT_BALANCE + CURRENT_BALANCE_DATE
# here -- the Summary tab and Reconciliation bridge both read from these two lines, so
# they can no longer silently go stale relative to each other.
# STATEMENT_* = the last FULLY reconciled official bank statement (all transactions matched, ties to the cent).
# Only update this pair when Kobi provides a fresh full statement export (not a spot balance).
STATEMENT_DATE = "2026-06-25"
STATEMENT_BALANCE = 35202.94
# CURRENT_* = the most recent balance figure reported, even if just a spot-check (not a full statement).
CURRENT_BALANCE_DATE = "2026-07-13"
CURRENT_BALANCE = 70000.00

# ============ RENOVATION LEDGER ============
# (date, category, description, vendor, amount, method, nbg, status)
L=[
("2026-01-17","Waste/Bins","Recycle/waste bins (8)","Kairis",992,"cash","Yes","Paid","2026-01-17"),
("2026-01-17","Labour","Team leader/manager (11 days)","Meli",880,"cash","Yes","Paid","2026-01-17"),
("2026-01-17","Labour","Workers (18 worker-days)","Workers",900,"cash","Yes","Paid","2026-01-17"),
("2026-01-17","Electrical","Electricity materials","Supplier",250,"cash","Yes","Paid","2026-01-17"),
("2026-01-17","Equipment","Tool damage replacement","-",250,"cash","Yes","Paid","2026-01-17"),
("2026-01-24","Waste/Bins","Recycle/waste bins (2)","Kairis",248,"cash","Yes","Paid","2026-01-24"),
("2026-01-24","Materials","Cement - large storage base","Amanatidis",431.48,"NBG card","Yes","Paid","2026-01-24"),
("2026-01-24","Labour","Workers (16 worker-days)","Workers",800,"cash","Yes","Paid","2026-01-24"),
("2026-01-24","Labour","Team leader (7 days)","Meli",560,"cash","Yes","Paid","2026-01-24"),
("2026-01-24","Equipment","Tool spare parts/materials","Poutichidis",120,"cash","Yes","Paid","2026-01-24"),
("2026-01-24","Materials","Steel - large storage base","Kairis",350,"cash","Yes","Paid","2026-01-24"),
("2026-02-02","Materials","Steel reinforcements","Kairis",220,"cash","Yes","Paid","2026-02-02"),
("2026-02-02","Materials","Cement - large storage slab","Amanatidis",800.07,"NBG card","Yes","Paid","2026-02-02"),
("2026-02-02","Labour","Team leader (5 days)","Meli",400,"cash","Yes","Paid","2026-02-02"),
("2026-02-02","Labour","Workers (7 worker-days)","Workers",350,"cash","Yes","Paid","2026-02-02"),
("2026-02-02","Waste/Bins","Recycle/waste bins (2)","Kairis",248,"cash","Yes","Paid","2026-02-02"),
("2026-02-02","Materials","Steel - large storage base","Kairis",350,"cash","Yes","Paid","2026-02-02"),
("2026-02-07","Other","Soula - inconvenience & water","Soula",50,"cash","Yes","Paid","2026-02-07"),
("2026-02-07","Labour","Workers (9 worker-days)","Workers",450,"cash","Yes","Paid","2026-02-07"),
("2026-02-07","Labour","Team leader (5 days)","Meli",400,"cash","Yes","Paid","2026-02-07"),
("2026-02-07","Waste/Bins","Recycle/waste bins (1)","Kairis",124,"cash","Yes","Paid","2026-02-07"),
("2026-02-07","Fuel","Fuel - concrete mixer (storage slab)","Gas station",30,"cash","Yes","Paid","2026-02-07"),
("2026-03-16","Labour","Workers (12 wd) - storage/road/trees/elec","Workers",600,"cash","Yes","Paid","2026-03-16"),
("2026-03-16","Labour","Team leader (3 days)","Meli",240,"cash","Yes","Paid","2026-03-16"),
("2026-03-16","Legal/Admin","Municipality road-barrier removal","Municipality",350,"cash","Yes","Paid","2026-03-16"),
("2026-03-16","Electrical","Electrician - bury backyard line + materials","Electrician",600,"cash","Yes","Paid","2026-03-16"),
("2026-03-16","Site Works","Excavator + 8 gravel trucks (160m road)","Alexiou",3700,"cash","Yes","Paid","2026-03-16"),
("2026-03-16","Materials","Kairis - Ytong + glue (storage walls)","Kairis",2216,"cash","Yes","Paid","2026-03-16"),
("2026-05-07","Legal/Admin","Workers insurance Jan-Feb (60 work-days)","EFKA",2200,"cash","Yes","Paid","2026-05-07"),
("2026-03-21","Materials","Kairis - roof insul/steel/fiberboard (adv.)","Kairis",500,"cash","Yes","Paid","2026-03-21"),
("2026-03-21","Labour","Team leader (5 days)","Meli",400,"cash","Yes","Paid","2026-03-21"),
("2026-03-21","Labour","Workers (6 worker-days)","Workers",300,"cash","Yes","Paid","2026-03-21"),
("2026-03-21","Legal/Admin","Greek SIM (Kobi)","Zoe",50,"cash","Yes","Paid","2026-03-21"),
("2026-03-21","Materials","Glass bricks - storage","Supplier",75,"cash","Yes","Paid","2026-03-21"),
("2026-03-21","Landscaping","Gardener (Mar-Apr)","Gardener",200,"cash","Yes","Paid","2026-03-21"),
("2026-03-28","Materials","Kairis - internal walls 1F (bricks/steel/sand)","Kairis",2950,"cash","Yes","Paid","2026-03-28"),
("2026-03-28","Labour","Team leader (3 days)","Meli",240,"cash","Yes","Paid","2026-03-28"),
("2026-03-28","Labour","Workers (6 worker-days)","Workers",300,"cash","Yes","Paid","2026-03-28"),
("2026-04-04","Labour","Brick workers - first floor","Workers",930,"cash","Yes","Paid","2026-04-04"),
("2026-04-07","Landscaping","Gardener pipes/hoses","Gardener",100,"cash","Yes","Paid","2026-04-07"),
("2026-04-07","Roof","Kairis - large storage roof materials","Kairis",1350,"cash","Yes","Paid","2026-04-07"),
("2026-04-07","Plumbing","Plumbing pipes floor1+storage (card part)","Supplier",1140,"NBG card","Yes","Paid","2026-04-07"),
("2026-04-07","Plumbing","Plumbing pipes floor1+storage (cash part)","Supplier",660,"cash","Yes","Paid","2026-04-07"),
("2026-04-09","Plumbing","Plumber advance #1","Plumber",1000,"cash","Yes","Paid","2026-04-09"),
("2026-04-21","Plumbing","Additional plumbing materials","Supplier",660,"cash","Yes","Paid","2026-04-21"),
("2026-04-21","Materials","Storage insulation materials","Kairis",1600,"cash","Yes","Paid","2026-04-21"),
("2026-04-28","Labour","Team leader (8 days) - storage+house","Meli",640,"cash","Yes","Paid","2026-04-28"),
("2026-04-28","Labour","Workers (13 worker-days)","Workers",650,"cash","Yes","Paid","2026-04-28"),
("2026-04-30","Plumbing","Plumber advance #2 - additional work","Plumber",1000,"cash","Yes","Paid","2026-04-30"),
("2026-05-10","Labour","Team leader (8 days) - back gate/small storage","Meli",640,"cash","Yes","Paid","2026-05-10"),
("2026-05-10","Labour","Workers (16 worker-days)","Workers",800,"cash","Yes","Paid","2026-05-10"),
("2026-05-10","Fuel","Fuel - cement mixer","Gas station",30,"cash","Yes","Paid","2026-05-10"),
("2026-05-10","Windows/Iron","Advance - ironworker (large storage door)","Blacksmith",350,"cash","Yes","Paid","2026-05-10"),
("2026-05-10","Legal/Admin","Workers social security (remaining)","EFKA",200,"cash","Yes","Paid","2026-05-10"),
("2026-05-10","Materials","Kairis - forms/iron/cement/coat/lime","Kairis",1750,"cash","Yes","Paid","2026-05-10"),
("2026-05-09","Materials","Amanatidis - extra materials","Amanatidis",205.75,"NBG card","Yes","Paid","2026-05-09"),
("2026-03-23","Materials","Amanatidis - materials","Amanatidis",87.02,"NBG card","Yes","Paid","2026-03-23"),
("2026-05-16","Windows/Iron","Old front gate repair + wheels","Blacksmith",50,"cash","Yes","Paid","2026-05-16"),
("2026-05-16","Plumbing","Plumber payment #3","Plumber",500,"cash","Yes","Paid","2026-05-16"),
("2026-05-16","Roof","Kairis - 3F panels+small-storage roof+Ytong+coat","Kairis",2875,"cash","Yes","Paid","2026-05-16"),
("2026-05-16","Electrical","Electrician advance - Apt B conduits","Electrician",400,"cash","Yes","Paid","2026-05-16"),
("2026-05-16","Labour","Team leader (6 days)","Meli",480,"cash","Yes","Paid","2026-05-16"),
("2026-05-16","Labour","Workers (16 worker-days)","Workers",800,"cash","Yes","Paid","2026-05-16"),
("2026-05-23","Windows/Iron","Aluminium inner door+window - large storage (TECHNICAL)","TECHNICAL",1150,"NBG card","Yes","Paid","2026-05-23"),
("2026-05-23","Labour","Workers (18 wd) week May19-23","Workers",900,"cash","Yes","Paid","2026-05-23"),
("2026-05-23","Labour","Team leader Meli (4 days)","Meli",320,"cash","Yes","Paid","2026-05-23"),
("2026-05-23","Labour","Replacement contractor (4 days)","Contractor",400,"cash","Yes","Paid","2026-05-23"),
("2026-05-23","Waste/Bins","Kairis - waste containers (3)","Kairis",360,"cash","Yes","Paid","2026-05-23"),
("2026-05-23","Materials","Kairis - upper floor materials","Kairis",100,"cash","Yes","Paid","2026-05-23"),
("2026-05-26","Materials","Kairis - partial debt settlement (of €3,335)","Kairis",2330,"cash","Yes","Paid","2026-05-26"),
("2026-05-26","Legal/Admin","Owner certificates x2 (Yael+Kobi)","Vasiliki",200,"transfer (NBG)","Yes","Paid","2026-05-26"),
("2026-05-18","Legal/Admin","Tax registration - accountant (SIANOS)","Accountant",711,"NBG card","Yes","Paid","2026-05-18"),
("2026-05-06","Legal/Admin","Land registry - temporal certification","Land registry",100,"transfer (NBG)","Yes","Paid","2026-05-06"),
("2026-05-25","Labour","Ice cream - worker treat","Meli Ice Cream",50,"cash","Yes","Paid","2026-05-25"),
("2026-05-28","Site Works","Excavator - open + close plumbing/electrical channels","Excavator",600,"cash","Yes","Paid","2026-05-28"),
("2026-05-28","Plumbing","Plumber - 3 days corner-to-corner pipe sealing","Plumber",1000,"cash","Yes","Paid","2026-05-28"),
("2026-05-28","Electrical","Electrician advance - power to Yael unit","Electrician",1000,"cash","Yes","Paid","2026-05-28"),
("2026-05-28","Plumbing","Plumber - materials & pipes (infrastructure)","Plumber",300,"cash","Yes","Paid","2026-05-28"),
("2026-06-01","Labour","Workers wages week May26-31","Workers",1500,"cash","Yes","Paid","2026-06-01"),
("2026-06-01","Materials","3-component epoxy coating - Yael unit floor","Supplier",400,"cash","Yes","Paid","2026-06-01"),
("2026-06-06","Labour","Workers wages week Jun2-6","Workers+Meli",1780,"cash","Yes","Paid","2026-06-06"),
("2026-06-06","Materials","Precast septic rings+cover+crane - Yael unit","Supplier",650,"cash","Yes","Paid","2026-06-06"),
("2026-06-06","Waste/Bins","Kairis - waste container","Kairis",125,"cash","Yes","Paid","2026-06-06"),
("2026-06-06","Fuel","Site fuel","Gas station",50,"cash","Yes","Paid","2026-06-06"),
("2026-06-06","Other","Old mattress removal - front area","Removal",50,"cash","Yes","Paid","2026-06-06"),
("2026-06-06","Materials","Kairis - 10 sacks (running tab)","Kairis",105,"cash","Yes","Paid","2026-06-06"),
("2026-06-06","Materials","Kairis - partial payment (of €4,045)","Kairis",2000,"cash","Yes","Paid","2026-06-06"),
("2026-06-06","Plumbing","Bathroom fixtures - Yael unit","Diamantopoulos",614,"Wise","No","Paid","2026-06-06"),
("2026-06-07","Labour","Workers Jun7 Sunday (Meli+1)","Workers",130,"cash","Yes","Paid","2026-06-07"),
("2026-06-13","Plumbing","Transport - bathroom supplies from Athens","Istiaia transport",45,"cash","Yes","Paid","2026-06-13"),
("2026-06-13","Labour","Workers Mon+Tue (4/day)","Workers",400,"cash","Yes","Paid","2026-06-13"),
("2026-06-13","Labour","Meli Mon+Tue","Meli",160,"cash","Yes","Paid","2026-06-13"),
("2026-06-13","Labour","Workers Wed+Thu (4+ESAT)","Workers",500,"cash","Yes","Paid","2026-06-13"),
("2026-06-13","Labour","Meli Wed+Thu","Meli",160,"cash","Yes","Paid","2026-06-13"),
("2026-06-13","Labour","Workers Fri+Sat (5+ESAT)","Workers",600,"cash","Yes","Paid","2026-06-13"),
("2026-06-13","Labour","Meli Fri+Sat","Meli",160,"cash","Yes","Paid","2026-06-13"),
("2026-06-13","Labour","Ilir - roof welding (Fri)","Ilir",300,"cash","Yes","Paid","2026-06-13"),
("2026-06-13","Labour","Ilir - roof welding (Sat)","Ilir",300,"cash","Yes","Paid","2026-06-13"),
("2026-06-13","Fuel","Site fuel week Jun8-13","Gas station",40,"cash","Yes","Paid","2026-06-13"),
("2026-06-13","Materials","Kairis materials week Jun8-13","Kairis",450,"cash","Yes","Paid","2026-06-13"),
("2026-06-13","Electrical","Extension cable","Supplier",75,"cash","Yes","Paid","2026-06-13"),
("2026-06-13","Materials","Tarpaulin (rain protection)","Supplier",75,"cash","Yes","Paid","2026-06-13"),
("2026-06-13","Materials","Concrete roof anchors + resin","Supplier",80,"cash","Yes","Paid","2026-06-13"),
("2026-06-15","Materials","Kairis - partial payment (bal -> ~€400)","Kairis",2000,"cash","Yes","Paid","2026-06-15"),
("2026-06-18","Materials","Athens shopping - Yael unit fit-out (door/AC/heater/window/access.)","Athens suppliers",972.55,"card (Vicky)","No","Paid","2026-06-18"),
("2026-06-24","Equipment","Demolition hammer drill (STRIGOS) - charge to Antonis contract","STRIGOS SERVICE",380.60,"NBG card","Yes","Paid","2026-06-24"),
("2026-06-19","Tiles","Spanish tiles 'Crosscut Marfil' 60x60 25.88m²","Tsintonis",286,"Wise","No","Paid","2026-06-19"),
("2026-06-22","Labour","Workers week Jun16-21 (incl polished cement)","Workers+Meli",2280,"cash","Yes","Paid","2026-06-22"),
("2026-06-22","Materials","Kairis materials week Jun15-21","Kairis",1175,"cash","Yes","Paid","2026-06-22"),
("2026-06-22","Furniture","Iron bed frame+bedside tables+slats - Yael","Supplier",350,"cash","Yes","Paid","2026-06-22"),
("2026-06-22","Furniture","Bedding/pillows/bedspread - Yael unit","Vicky",155,"card (Vicky)","No","Paid","2026-06-22"),
("2026-06-22","Fuel","Site fuel week Jun16-21","Gas station",40,"cash","Yes","Paid","2026-06-22"),
("2026-06-24","Materials","Kairis debt settlement (old bal+materials)","Kairis",1800,"cash","Yes","Paid","2026-06-24"),
("2026-06-24","Materials","Beam (structural) + rendering materials","Kairis",420,"cash","Yes","Paid","2026-06-24"),
("2026-06-25","Furniture","Custom mattress - Yael unit (195x170 30cm)","Local mfr",400,"cash","Yes","Paid","2026-06-25"),
("2026-06-27","Materials","Cleaning materials - Yael unit (Jun 27)","Supplier",150.00,"cash","Yes","Paid","2026-06-27"),
("2026-07-05","Materials","Sand and gravel - week Jun29-Jul5","Supplier",150.00,"cash","Yes","Paid","2026-07-10"),
("2026-07-05","Fuel","Fuel - concrete mixers week Jun29-Jul5","Gas station",40.00,"cash","Yes","Paid","2026-07-10"),
("2026-07-05","Materials","Steel reinforcement - wall bond beams above door openings","Supplier",400.00,"cash","Yes","Paid","2026-07-10"),
("2026-07-05","Materials","Kairis - column reinforcement steel, cement, resin anchors","Kairis",1800.00,"cash","Yes","Paid","2026-07-10"),
("2026-07-05","Labour","Workers wages - week Jun29-Jul5","Workers",1910.00,"cash","Yes","Paid","2026-07-10"),
("2026-07-10","Materials","Kairis - SikaGrout/reinforcement materials (partial payment toward new balance)","Kairis",1700.00,"cash","Yes","Paid","2026-07-10"),
("2026-07-12","Labour","Workers wages - week Jul7-12","Workers",2150.00,"cash","Yes","Paid","2026-07-12"),
("2026-07-19","Labour","Workers wages Mon-Wed - week Jul13-19","Workers",900.00,"cash","Yes","Paid","2026-07-20"),
("2026-07-19","Equipment","Concrete mixer - week Jul13-19","Site",20.00,"cash","Yes","Paid","2026-07-20"),
("2026-07-19","Materials","Binding wire and steel-tying materials - week Jul13-19","Supplier",65.00,"cash","Yes","Paid","2026-07-20"),
# committed / upcoming (approved)
("2026-06-28","Materials","Materials - week Jun25-28 (Yael unit completion)","Kairis",1500.00,"cash","Yes","Paid","2026-06-29"),
("2026-06-28","Plumbing","Plumber - Yael unit completion","Plumber",200.00,"cash","Yes","Paid","2026-06-29"),
("2026-06-28","Labour","Carpenter - Yael unit (bedside tables/countertop/door)","Carpenter",250.00,"cash","Yes","Paid","2026-06-29"),
("2026-06-28","Materials","Marble for shower - Yael unit","Supplier",50.00,"cash","Yes","Paid","2026-06-29"),
("2026-06-28","Fuel","Site fuel week Jun25-28","Vicky",40.00,"cash","Yes","Paid","2026-06-29"),
("2026-06-28","Labour","Workers week Jun25-28 (Yael unit + site)","Workers",2600.00,"cash","Yes","Paid","2026-06-29"),
("2026-05-23","Electrical","Electrician - 2 apts+2 storages+outdoor","Electrician",9500,"cash","Yes","Committed",""),
("2026-05-23","Electrical","Electrical plans - submitted to authority","Electrician",1000,"-","Yes","Committed",""),
("2026-05-23","Electrical","Three-phase meter (Kobi's name)","Electrician",750,"-","Yes","Committed",""),
("2026-05-23","Electrical","Electrician advance - 130m piping reimb.","Electrician",1000,"cash","Yes","Committed",""),
("2026-05-23","Windows/Iron","Metalworker - small storage door+window","Metalworker",550,"cash","Yes","Committed",""),
("2026-06-15","Site Works","Excavator - road compaction (60t roller)","Excavator",600,"cash","Yes","Committed",""),
("2026-07-20","Legal/Admin","Worker insurance Apr-Jun - partial payment","EFKA",1865,"cash","Yes","Paid","2026-07-20"),
("2026-06-25","Legal/Admin","Worker insurance Apr-Jun - remaining balance","EFKA",1993,"-","Yes","Committed",""),
("2026-07-19","Materials","Kairis - partial payment (ATM shortfall from last week)","Kairis",150,"cash","Yes","Paid","2026-07-20"),
("2026-06-24","Materials","Kairis - remaining balance","Kairis",45,"cash","Yes","Committed",""),
("2026-07-11","Materials","Kairis - next SikaGrout order (200 bags) + 20 steel bars - approved by Yael, not yet placed","Kairis",7740.00,"cash","Yes","Committed",""),
("2026-07-11","Materials","Kairis - outstanding debt for column/beam materials received this week (EUR3,420+EUR3,530=EUR6,940, minus EUR1,700 already paid Jul10)","Kairis",5240.00,"cash","Yes","Committed",""),
("2026-07-11","Materials","Kairis - additional item per Vicky Jul11 running-total correction (detail unspecified, confirms total EUR13,730)","Kairis",750.00,"cash","Yes","Committed",""),
]

ws=wb.active; ws.title="Renovation Ledger"
title(ws,"Renovation Ledger — Ag. Nikolaos Home","Every renovation expense (Jan–Jun 2026), de-duplicated. The €380 demolition hammer drill (18-Jun) is a SEPARATE payment, paid by Vicky's card, to be deducted from the Antonis contract.",8)
hdr(ws,4,["Report Date","Category","Description","Vendor / Payee","Amount (€)","Method","From NBG?","Status","Payment Date"])
r=5
for d,cat,desc,ven,amt,meth,nbg,st,pd in L:
    ws.cell(r,1,d); ws.cell(r,2,cat); ws.cell(r,3,desc); ws.cell(r,4,ven)
    c=ws.cell(r,5,amt); c.number_format=EUR; c.alignment=right
    ws.cell(r,6,meth); ws.cell(r,7,nbg); ws.cell(r,8,st); ws.cell(r,9,pd)
    fc = GREEN if st=="Paid" else YELL
    ws.cell(r,8).fill=fill(fc)
    for j in range(1,10): ws.cell(r,j).font=F(9); ws.cell(r,j).border=border; ws.cell(r,j).alignment=(right if j==5 else (center if j in(6,7,8) else left))
    r+=1
last=r-1
# totals
ws.cell(r,4,"PAID — total").font=F(10,True); ws.cell(r,4).alignment=right
c=ws.cell(r,5,f'=SUMIF(H5:H{last},"Paid",E5:E{last})'); c.font=F(10,True); c.number_format=EUR; c.fill=fill(LBLUE)
r+=1
ws.cell(r,4,"COMMITTED (approved, unpaid)").font=F(10,True); ws.cell(r,4).alignment=right
c=ws.cell(r,5,f'=SUMIF(H5:H{last},"Committed",E5:E{last})'); c.font=F(10,True); c.number_format=EUR; c.fill=fill(YELL)
r+=1
ws.cell(r,4,"PAID — from NBG (cash/transfer)").font=F(9,True); ws.cell(r,4).alignment=right
c=ws.cell(r,5,f'=SUMIFS(E5:E{last},H5:H{last},"Paid",G5:G{last},"Yes")'); c.font=F(9,True); c.number_format=EUR
r+=1
ws.cell(r,4,"PAID — outside NBG (Wise/card)").font=F(9,True); ws.cell(r,4).alignment=right
c=ws.cell(r,5,f'=SUMIFS(E5:E{last},H5:H{last},"Paid",G5:G{last},"No")'); c.font=F(9,True); c.number_format=EUR
PAID_CELL=f"'Renovation Ledger'!E{last+1}"; COMMIT_CELL=f"'Renovation Ledger'!E{last+2}"; NBGPAID_CELL=f"'Renovation Ledger'!E{last+3}"
for col,w in zip("ABCDEFGHI",[13,13,42,20,13,12,9,12,13]): ws.column_dimensions[col].width=w
ws.freeze_panes="A5"

# ============ FUNDING (DEPOSITS) ============
ws2=wb.create_sheet("Funding (NBG Deposits)")
title(ws2,"Funding — Deposits into NBG Account","Israel → NBG GR18 0110 3180 0000 3180 0400 611. All amounts EUR received. Source: bank transfer confirmations in /Money Transfers.",5)
hdr(ws2,4,["Value date","Amount (€)","Bank ref","Greek-side confirmation","Counted?"])
DEP=[("2025-09-28",25000,"55423","PDF OtherBankMatachTransfer_25_09 to GR18-0400611","Yes"),
("2025-09-28",15000,"55758","PDF OtherBankMatachTransfer_26_09","Yes"),
("2025-10-12",130000,"63360","PDF OtherBankMatachTransfer_12_10","Yes"),
("2026-01-11",10000,"14452","PDF OtherBankMatachTransfer_10_01","Yes"),
("2026-02-08",20000,"30860","Confirmed - bank statement (ref 30860) + project_data.js","Yes"),
("2026-04-26",70000,"70724","PDF OtherBankMatachTransfer_26_04","Yes"),
("2026-05-31",2000,"90833","Transfer to WISE as backup - not spent; NOT a Greek NBG deposit","No"),
("2026-07-13",50000,"wire-Jul13","Kobi confirmed - Israel to NBG transfer (Mizrahi)","Yes"),]
r=5
for d,amt,ref,conf,cnt in DEP:
    ws2.cell(r,1,d);c=ws2.cell(r,2,amt);c.number_format=EUR0;c.alignment=right;ws2.cell(r,3,ref);ws2.cell(r,4,conf);cc=ws2.cell(r,5,cnt)
    cc.fill=fill(GREEN if cnt=="Yes" else YELL)
    for j in range(1,6): ws2.cell(r,j).font=F(9);ws2.cell(r,j).border=border;ws2.cell(r,j).alignment=(right if j==2 else (center if j in(1,5) else left))
    r+=1
dl=r-1
ws2.cell(r,1,"TOTAL DEPOSITS (counted)").font=F(11,True);ws2.cell(r,1).alignment=right
c=ws2.cell(r,2,f'=SUMIF(E5:E{dl},"Yes",B5:B{dl})');c.font=F(11,True);c.number_format=EUR0;c.fill=fill(LBLUE)
DEP_CELL=f"'Funding (NBG Deposits)'!B{r}"
r+=2
ws2.cell(r,1,"Ground truth: Mizrahi-Tefahot euro account statement (Sikum Haavarot Matach 2.pdf, 01/01/2024-22/06/2026). Rows above are the transfer-to-another-bank lines to the Greek NBG account in the project period. 2024 transfers (EUR 10k, 1.7k) excluded as pre-project.").font=F(8,False,"595959")
ws2.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5);r+=1
ws2.cell(r,1,"Israeli euro account still held ~EUR 64,786 liquid + EUR 70,000 in a term deposit (23-Apr) as of 10-Jun-2026 - available for future project funding.").font=F(8,False,"595959")
ws2.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5)
for col,w in zip("ABCDE",[13,14,10,46,11]): ws2.column_dimensions[col].width=w

# ============ PROPERTY ACQUISITION ============
ws3=wb.create_sheet("Property Acquisition 2025")
title(ws3,"Property Acquisition (2025)","Purchase of two-storey residence + plot, Agios Nikolaos. Contract price €145,000 (objective value €239,793.82). Paid from NBG account.",5)
hdr(ws3,4,["Date","Item","Payee","Amount (€)","From NBG?","Note"])
ACQ=[("2025-09-29","House payment 1 - pre-agreement advance","Anagnostou",18500,"Yes","Bank transfer"),
("2025-10-22","House payment 2 - via Zoe (issued check to seller)","Anagnostou / Zoe",107000,"Yes","Bank transfer to Zoe -> check to Anagnostou"),
("2025-10-22","House payment 2 - balance (direct)","Anagnostou",19731,"Yes","Direct bank transfer"),
("2025-10-17","Purchase tax (FMA, incl 0.9% joint discount)","via Zoi Koliou",7410,"Yes","Transfer to Zoe"),
("2025-11-05","Zoe commission","Zoi Koliou",4350,"Yes","ATM cash via our debit card (800+2000+1000+550)"),
("2025-11-05","Attorney - Vasiliki","Vasiliki",2000,"Yes","Cash withdrawn by Zoe"),
("2025-11-05","Land registration fee","Vasiliki",1328,"Yes","Transfer to Vasiliki (pays on our behalf)"),
("2025-10-24","Notary (Vasiliki)","Mavri / Vasiliki",2490,"Yes","Direct bank transfer"),
("2025-09-29","Legalization (engineer) - transfer 1","Vicky (Revolut LT)",800,"Yes","Pergola/rooftop/2 warehouses/wells"),
("2025-10-01","Legalization (engineer) - transfer 2","Vicky (Revolut LT)",520,"Yes",""),
("2025-10-01","Legalization (engineer) - transfer 3","Antonis",680,"Yes",""),
("2025","Legalization (engineer) - cash balance","Cash",3000,"Yes","Cash portion (total legalization = EUR 5,000)"),
("2025","Notary & apostille - NBG account setup","-",2700,"No","Paid from Israel (to open the account)"),
("2025","Flight to Evia - purchase trip","-",1200,"No","Travel, paid from Israel"),]
r=5
for d,it,py,amt,nbg,note in ACQ:
    ws3.cell(r,1,d);ws3.cell(r,2,it);ws3.cell(r,3,py);c=ws3.cell(r,4,amt);c.number_format=EUR0;c.alignment=right
    cc=ws3.cell(r,5,nbg);cc.fill=fill(GREEN if nbg=="Yes" else YELL);ws3.cell(r,6,note)
    for j in range(1,7): ws3.cell(r,j).font=F(9);ws3.cell(r,j).border=border;ws3.cell(r,j).alignment=(right if j==4 else (center if j==5 else left))
    r+=1
al=r-1
ws3.cell(r,3,"ACQUISITION TOTAL (all)").font=F(11,True);ws3.cell(r,3).alignment=right
c=ws3.cell(r,4,f"=SUM(D5:D{al})");c.font=F(11,True);c.number_format=EUR0;c.fill=fill(LBLUE)
ACQ_CELL=f"'Property Acquisition 2025'!D{r}"
r+=1
ws3.cell(r,3,"of which paid from NBG").font=F(10,True);ws3.cell(r,3).alignment=right
c=ws3.cell(r,4,f'=SUMIF(E5:E{al},"Yes",D5:D{al})');c.font=F(10,True);c.number_format=EUR0;c.fill=fill(LBLUE)
ACQ_NBG_CELL=f"'Property Acquisition 2025'!D{r}"
r+=2
ws3.cell(r,1,"Source: 'Agio Nikolaos house payments' sheet, 'Actual' table. House price EUR 145,231 (18,500 + 107,000 + 19,731).").font=F(8,False,"595959")
ws3.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
for col,w in zip("ABCDEF",[12,42,22,13,10,44]): ws3.column_dimensions[col].width=w

# ============ OUT-OF-NBG & REIMBURSEMENTS ============
ws4=wb.create_sheet("Out-of-NBG & Reimburse")
title(ws4,"Payments Outside NBG  +  Vicky Reimbursements","Project costs paid via Wise or by card (not from the NBG account), plus reimbursement transfers to Vicky. These expenses are already counted once in the Renovation Ledger.",5)
hdr(ws4,4,["Date","Item","Method","Amount (€)","Note"])
OUT=[("2026-06-06","Bathroom fixtures (Diamantopoulos)","Wise",614,"In ledger"),
("2026-06-18","Athens shopping - Yael unit fit-out","card (Vicky)",972.55,"In ledger; reimbursed €972.55 below"),
("2026-06-19","Spanish tiles (Tsintonis)","Wise",286,"In ledger; NBG was blocked"),
("2026-06-22","Bedding/pillows/bedspread - Yael","card (Vicky)",155,"In ledger; to net vs window cost"),]
r=5
for d,it,me,amt,nt in OUT:
    ws4.cell(r,1,d);ws4.cell(r,2,it);ws4.cell(r,3,me);c=ws4.cell(r,4,amt);c.number_format=EUR;c.alignment=right;ws4.cell(r,5,nt)
    for j in range(1,6): ws4.cell(r,j).font=F(9);ws4.cell(r,j).border=border;ws4.cell(r,j).alignment=(right if j==4 else left)
    r+=1
ol=r-1
ws4.cell(r,3,"Subtotal").font=F(10,True);ws4.cell(r,3).alignment=right
c=ws4.cell(r,4,f"=SUM(D5:D{ol})");c.font=F(10,True);c.number_format=EUR;c.fill=fill(LBLUE)
r+=2
ws4.cell(r,1,"REIMBURSEMENTS TO VICKY (cash-flow only — not additional project cost)").font=F(10,True,"C00000")
ws4.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5);r+=1
hdr(ws4,r,["Date","Item","Method","Amount (€)","Note"]);r+=1
ws4.cell(r,1,"2026-06-22");ws4.cell(r,2,"Reimburse Athens/Praktiker card spend");ws4.cell(r,3,"Wise → Revolut")
c=ws4.cell(r,4,972.55);c.number_format=EUR;c.alignment=right
ws4.cell(r,5,"Settles only the €972.55 Praktiker spend (not the €380 drill). LT853250002013391261")
for j in range(1,6): ws4.cell(r,j).font=F(9);ws4.cell(r,j).border=border;ws4.cell(r,j).alignment=(right if j==4 else left)
for col,w in zip("ABCDE",[12,40,16,14,40]): ws4.column_dimensions[col].width=w

# ============ CATEGORY ROLLUP ============
ws5=wb.create_sheet("Category Rollup")
title(ws5,"Category Rollup — Paid vs Budget","Paid = actuals from the Renovation Ledger (SUMIF). Budget = working estimates (Budget Overview / project_data forecast). Remaining is indicative.",4)
hdr(ws5,4,["Category","Paid to date (€)","Working budget (€)","Notes"])
CATS=[("Labour",25000,"Workers, Meli, ESAT, Ilir, contractors"),
("Materials",25000,"Kairis main account, Amanatidis, sundries"),
("Plumbing",9000,"All floors; ~€7.2k balance after water connection"),
("Electrical",11250,"GF+storage approved €11,250; upper floors TBD"),
("Site Works",6000,"Excavation, 160m road, compaction"),
("Roof",4500,"3F panels, storage roofs"),
("Windows/Iron",18700,"Storage done; ~25 house openings upcoming"),
("Waste/Bins",2000,"Recycle/waste containers"),
("Tiles",25500,"€286 paid; ~€25k order+install upcoming"),
("Furniture",1000,"Yael unit bed, mattress, bedding"),
("Legal/Admin",11350,"Insurance, certificates, SIM, accountant"),
("Landscaping",5000,"Gardener, irrigation"),
("Fuel",500,"Site fuel"),
("Equipment",500,"Tools"),
("Other",300,"Misc"),]
r=5; ll=5
for cat,bud,note in CATS:
    ws5.cell(r,1,cat)
    c=ws5.cell(r,2,f"=SUMIFS('Renovation Ledger'!E5:E{last},'Renovation Ledger'!B5:B{last},A{r},'Renovation Ledger'!H5:H{last},\"Paid\")");c.number_format=EUR;c.alignment=right
    c2=ws5.cell(r,3,bud);c2.number_format=EUR0;c2.alignment=right
    ws5.cell(r,4,note)
    for j in range(1,5): ws5.cell(r,j).font=F(9);ws5.cell(r,j).border=border;ws5.cell(r,j).alignment=(right if j in(2,3) else left)
    r+=1
le=r-1
ws5.cell(r,1,"TOTAL").font=F(10,True)
c=ws5.cell(r,2,f"=SUM(B{ll}:B{le})");c.font=F(10,True);c.number_format=EUR;c.fill=fill(LBLUE)
c=ws5.cell(r,3,f"=SUM(C{ll}:C{le})");c.font=F(10,True);c.number_format=EUR0;c.fill=fill(LBLUE)
for col,w in zip("ABCD",[16,18,18,46]): ws5.column_dimensions[col].width=w

# ============ RECONCILIATION ============
ws6=wb.create_sheet("Reconciliation")
title(ws6,"NBG Account Reconciliation (official statement)","Account 31800400611, period 01/10/2025-25/06/2026. Ties to the cent. Source: NBG statement export25-06-2026.xlsx.",4)
rows=[
("Opening balance 01/10/2025",21066.81,LBLUE,"Embeds Sep-2025: +EUR25k +EUR15k transfers, -EUR18,500 down-payment, -fees"),
("+ Deposits received (in period)",230000,GREEN,"EUR130k (13-Oct) + EUR10k (12-Jan) + EUR20k (09-Feb) + EUR70k (27-Apr)"),
("- Property acquisition & legal (transfers)",-139459.28,None,"House 126,730.81 + tax 7,410 + notary 2,490 + land reg 1,328.47 + owner cert 200 + land-reg temp 100 + legalization 620 + 580"),
("- Acquisition paid in cash (Nov-2025)",-6350,None,"Zoe commission 4,350 + Vasiliki 2,000"),
("- Renovation/project cash withdrawals (ATM)",-64730,None,"Cash for workers, Kairis, materials, etc."),
("- Renovation card payments (POS)",-4494.92,None,"Amanatidis 1,524.32 + GKIZA 1,440 + aluminium 1,150 + drill 380.60"),
("- Accountant - tax registration",-711,None,"SIANOS PANTAZIS, 18-May"),
("- Personal card spend",-104.17,None,"Jan groceries/cafe (LIDL etc.)"),
("- Bank fees & charges",-14.50,None,"In-period commissions"),
]
r=5
for lbl,val,fc,note in rows:
    ws6.cell(r,2,lbl).font=F(10, lbl.startswith("+") or lbl.startswith("Opening"))
    c=ws6.cell(r,3,val);c.number_format=EUR;c.alignment=right;c.font=F(10)
    if fc: c.fill=fill(fc); ws6.cell(r,2).fill=fill(fc)
    ws6.cell(r,4,note).font=F(8,False,"595959")
    ws6.cell(r,2).border=border;c.border=border
    r+=1
ws6.cell(r,2,"Computed closing balance").font=F(10,True)
c=ws6.cell(r,3,f"=SUM(C5:C{r-1})");c.font=F(10,True);c.number_format=EUR;c.fill=fill(LBLUE);c.alignment=right;c.border=border
r+=1
ws6.cell(r,2,"Statement closing balance 25/06/2026").font=F(10,True)
c=ws6.cell(r,3,STATEMENT_BALANCE);c.font=F(10,True);c.number_format=EUR;c.fill=fill(YELL);c.alignment=right;c.border=border
r+=1
ws6.cell(r,2,"Difference (ties to zero)").font=F(11,True)
c=ws6.cell(r,3,"=C14-C15");c.font=F(11,True);c.number_format=EUR;c.fill=fill(GREEN);c.alignment=right;c.border=border
r+=2
ws6.cell(r,2,"Cross-check - renovation funded from NBG").font=F(10,True);r+=1
ws6.cell(r,2,"Renovation from NBG per ledger").font=F(9)
c=ws6.cell(r,3,f"={NBGPAID_CELL}");c.number_format=EUR;c.font=F(9);c.alignment=right;r+=1
ws6.cell(r,2,"Statement NBG renovation outflow (cash 64,730 + POS 4,494.92 + accountant 711)").font=F(9)
c=ws6.cell(r,3,69935.92);c.number_format=EUR;c.font=F(9);c.alignment=right;r+=1
ws6.cell(r,2,"Difference = cash float / timing (+ ~300 legal items the statement groups under acquisition)").font=F(9,True)
c=ws6.cell(r,3,f"={NBGPAID_CELL}-69935.92");c.number_format=EUR;c.font=F(9,True);c.alignment=right;r+=2
for n in ["The NBG account is fully reconciled to the official statement - there is NO unexplained money.",
"The small ledger-vs-statement difference is cash float (cash withdrawn vs spent timing). ATM withdrawals are exact; renovation cash expenses are reconstructed weekly estimates."]:
    ws6.cell(r,2,n).font=F(9);ws6.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4);ws6.cell(r,2).alignment=left;r+=1

r+=1
ws6.cell(r,2,f"ROLLING BRIDGE — since last full statement ({STATEMENT_DATE})").font=F(11,True,WHITE);ws6.cell(r,2).fill=fill(BLUE)
ws6.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4);r+=1
ws6.cell(r,2,f"Opening — statement closing balance ({STATEMENT_DATE})").font=F(9)
c=ws6.cell(r,3,STATEMENT_BALANCE);c.number_format=EUR;c.font=F(9);c.alignment=right
opening_row=r; r+=1
ws6.cell(r,2,"+ Deposits since (dynamic — recalculates as new transfers are added)").font=F(9)
c=ws6.cell(r,3,f"=SUMPRODUCT(('Funding (NBG Deposits)'!A5:A{dl}>\"{STATEMENT_DATE}\")*('Funding (NBG Deposits)'!E5:E{dl}=\"Yes\")*'Funding (NBG Deposits)'!B5:B{dl})")
c.number_format=EUR;c.font=F(9);c.alignment=right
dep_row=r; r+=1
ws6.cell(r,2,"- Renovation paid from NBG since (dynamic — recalculates as new expenses are added)").font=F(9)
c=ws6.cell(r,3,f"=SUMPRODUCT(('Renovation Ledger'!I5:I{last}>\"{STATEMENT_DATE}\")*('Renovation Ledger'!H5:H{last}=\"Paid\")*('Renovation Ledger'!G5:G{last}=\"Yes\")*'Renovation Ledger'!E5:E{last})")
c.number_format=EUR;c.font=F(9);c.alignment=right
out_row=r; r+=1
ws6.cell(r,2,"Computed balance (bridge)").font=F(10,True)
c=ws6.cell(r,3,f"=C{opening_row}+C{dep_row}-C{out_row}");c.font=F(10,True);c.number_format=EUR;c.fill=fill(LBLUE);c.alignment=right
computed_row=r; r+=1
ws6.cell(r,2,f"Reported balance ({CURRENT_BALANCE_DATE})").font=F(10,True)
c=ws6.cell(r,3,CURRENT_BALANCE);c.font=F(10,True);c.number_format=EUR;c.fill=fill(YELL);c.alignment=right
reported_row=r; r+=1
ws6.cell(r,2,"Residual (bridge − reported)").font=F(11,True)
_residual = STATEMENT_BALANCE - CURRENT_BALANCE  # deposits since = 0 currently; adjust if that changes
c=ws6.cell(r,3,f"=C{computed_row}-C{reported_row}");c.font=F(11,True);c.number_format=EUR
c.fill=fill(GREEN if abs(_residual)<50 else RED);c.alignment=right
r+=1
ws6.cell(r,2,"A residual here is expected cash-float/timing (ATM withdrawals vs itemized weekly reconstructions), or spending not yet reported to Kobi -- see Flags & Dedup. It is NOT unexplained/missing money unless it grows unexpectedly large or keeps widening.").font=F(8,False,"595959")
ws6.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4);r+=1

for col,w in zip("ABCD",[3,46,16,54]): ws6.column_dimensions[col].width=w

# ============ SUMMARY ============
ws0=wb.create_sheet("Summary"); wb.move_sheet("Summary",-(len(wb.sheetnames)-1))
title(ws0,"Ag. Nikolaos Home — Project Balance Sheet",f"Financial ground truth. Currency: EUR. Last updated {CURRENT_BALANCE_DATE}. Sources: project folder, Drive payments sheet, WhatsApp 'Evia HOME', Gmail, bank confirmations.",4)
def kpi(ws,row,label,formula,fc=LBLUE,fmt=EUR0,note=""):
    ws.cell(row,1,label).font=F(11,True);ws.cell(row,1).alignment=left
    c=ws.cell(row,2,formula);c.font=F(12,True,NAVY);c.number_format=fmt;c.fill=fill(fc);c.alignment=center;c.border=border
    ws.cell(row,3,note).font=F(9,False,"595959");ws.cell(row,3).alignment=left
r=4
ws0.cell(r,1,"FUNDING & POSITION").font=F(12,True,WHITE);ws0.cell(r,1).fill=fill(BLUE)
ws0.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4);r+=1
kpi(ws0,r,"Total deposited into NBG",f"={DEP_CELL}",LBLUE,EUR0,"6 transfers Sep-2025 to May-2026 (bank statement)");r+=1
kpi(ws0,r,"Property acquisition (2025)",f"={ACQ_CELL}",GREY,EUR0,"Contract price €145,000 + legal/notary/tax");r+=1
kpi(ws0,r,"Current NBG balance",CURRENT_BALANCE,YELL,EUR0,f"Reported {CURRENT_BALANCE_DATE} (see Reconciliation > Rolling Bridge for the residual check)");r+=1
r+=1
ws0.cell(r,1,"RENOVATION").font=F(12,True,WHITE);ws0.cell(r,1).fill=fill(BLUE)
ws0.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4);r+=1
kpi(ws0,r,"Renovation paid to date",f"={PAID_CELL}",GREEN,EUR,"All methods, de-duplicated");r+=1
kpi(ws0,r,"Committed (approved, unpaid)",f"={COMMIT_CELL}",YELL,EUR,"Electrician, metalworker, insurance, etc.");r+=1
kpi(ws0,r,"Acquisition + Renovation paid",f"={ACQ_CELL}+{PAID_CELL}",LBLUE,EUR0,"Total cash deployed into the project to date");r+=1
r+=1
ws0.cell(r,1,"RECONCILIATION CHECK").font=F(12,True,WHITE);ws0.cell(r,1).fill=fill(BLUE)
ws0.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4);r+=1
kpi(ws0,r,"NBG reconciliation difference",f"=Reconciliation!C16",GREEN,EUR0,"Account ties to the official statement (zero)");r+=1
r+=1
ws0.cell(r,1,"⚑ Open items needing your confirmation are on the 'Flags & Dedup' tab.").font=F(10,True,"C00000")
ws0.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
for col,w in zip("ABCD",[34,20,50,4]): ws0.column_dimensions[col].width=w

# ============ FLAGS ============
wsf=wb.create_sheet("Flags & Dedup")
title(wsf,"Flags & De-duplication Log","Resolved items (green) and open questions (yellow). Nothing here is double-counted in the totals.",5)
hdr(wsf,4,["#","Issue","How handled","Status","Action needed"])
FL=[
("1","Feb-8-2026 €20,000 transfer","Confirmed in the Mizrahi euro-account statement (ref 30860) and project_data.js. Counted in deposits.","Resolved","None."),
("2","Hammer drill €380 (18-Jun)","Confirmed SEPARATE from the €972.55 Athens shopping; added as its own line, charged to the Antonis contract.","Resolved","Recover the €380 by deducting it from the Antonis contract balance."),
("3","Vicky reimbursement €972.55 (Wise)","A settlement of the €972.55 Athens expense; shown only on the Out-of-NBG tab, never double-counted.","Resolved","None - you confirmed it was sent."),
("4","May-28 cluster: excavator €600, plumber €1,000, electrician €1,000, plumber materials €300","All FOUR confirmed real (WhatsApp 28-May + dashboard) and added to the ledger.","Resolved","None."),
("5","Early items missing from Drive sheet: aluminium €1,150, plumber adv#1 €1,000, gate €50","Confirmed they belong; added to the ledger.","Resolved","None."),
("6","Acquisition detail and the €7,410 to Zoi","Itemised from the Actual table: €7,410 is the property purchase TAX (FMA); balance split €107,000 (Zoe check) + €19,731 direct.","Resolved","None."),
("10","€2,000 transfer 31-May (ref 90833)","Went to Wise as an unspent backup buffer; excluded from NBG deposits.","Resolved","None."),
("7","Reconciliation residual","RESOLVED with the official NBG statement: the account reconciles to the cent (€35,202.94). The earlier residual was an expenses-vs-deposits artifact, not missing money.","Resolved","None."),
("8","Amanatidis + plumbing card payments (€2,140)","RESOLVED per Kobi: paid with the NBG debit card; reclassified as From NBG = Yes.","Resolved","None."),
("9","Worker insurance Apr-Jun €3,858","Not present in the NBG statement -> confirms still UNPAID. Listed as Committed.","OPEN","Pay/confirm; then it moves to Paid."),
("11","EUR907.92 residual — rolling bridge (25-Jun to 11-Jul)","Kobi reported the current NBG balance as EUR23,505.02 on 11-Jul-2026. The ledger only accounts for EUR10,790 in NBG-paid cash-out since the 25-Jun statement (EUR35,202.94), projecting EUR24,412.94 -- EUR907.92 more than reported. Consistent with the known cash-float pattern (round ATM withdrawals vs itemized weekly reconstructions) but could also be spending Vicky hasn't reported yet.","OPEN","Ask Vicky to confirm — either an ATM withdrawal that ran ahead of itemized expenses, or a payment not yet reported in the WhatsApp group."),
]
r=5
for n,iss,hand,st,act in FL:
    wsf.cell(r,1,n);wsf.cell(r,2,iss);wsf.cell(r,3,hand);sc=wsf.cell(r,4,st);wsf.cell(r,5,act)
    sc.fill=fill(GREEN if st=="Resolved" else YELL)
    for j in range(1,6): wsf.cell(r,j).font=F(9);wsf.cell(r,j).border=border;wsf.cell(r,j).alignment=(center if j in(1,4) else left)
    r+=1
for col,w in zip("ABCDE",[4,34,46,10,40]): wsf.column_dimensions[col].width=w

import os
wb.save(os.path.join(os.path.dirname(os.path.abspath(__file__)),"Ag_Nikolaos_Balance_Sheet.xlsx"))
print("saved; sheets:",wb.sheetnames)
